import os
from datetime import datetime, timedelta
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.database import SessionLocal
from app.models.report import Report, ReportStatus
from app.models.tender import Tender
from app.models.proposal import TenderProposal
from app.models.supplier_rating import SupplierRating
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from celery import shared_task
from app.config import settings


UPLOAD_DIR = settings.UPLOAD_DIR
os.makedirs(UPLOAD_DIR, exist_ok=True)


@shared_task
def generate_report_task(report_id: int):
    db: Session = SessionLocal()
    try:
        report = db.query(Report).filter(Report.id == report_id).first()
        if not report:
            return {"status": "error", "message": "Report not found"}
        
        report.status = ReportStatus.GENERATING
        db.commit()

        result_file = None
        file_type = None

        if report.type in ["tender_summary", "monthly_tender_pdf", "budget_analytics"]:
            pdf_filename = f"report_{report_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
            pdf_path = os.path.join(UPLOAD_DIR, pdf_filename)
            
            doc = SimpleDocTemplate(pdf_path, pagesize=letter)
            styles = getSampleStyleSheet()
            elements = []

            # Title
            title = Paragraph(report.title, styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 12))

            if report.type == "tender_summary":
                tenders = db.query(Tender).all()
                data = [["ID", "Тақырыбы", "Статус", "Дедлайн"]]
                for tender in tenders:
                    data.append([tender.id, tender.title, tender.status.value, tender.deadline.strftime("%Y-%m-%d")])
                
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.grey),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0,0), (-1,0), 12),
                    ('BACKGROUND', (0,1), (-1,-1), colors.beige),
                    ('GRID', (0,0), (-1,-1), 1, colors.black),
                ]))
                elements.append(table)
            
            elif report.type == "monthly_tender_pdf":
                now = datetime.utcnow()
                first_day = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                tenders = db.query(Tender).filter(Tender.created_at >= first_day).all()
                data = [["ID", "Тақырыбы", "Статус", "Дедлайн", "Бюджет"]]
                for tender in tenders:
                    data.append([tender.id, tender.title, tender.status.value, tender.deadline.strftime("%Y-%m-%d"), f"{tender.budget or 0} ₸"])
                
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4CAF50")),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0,0), (-1,0), 12),
                    ('BACKGROUND', (0,1), (-1,-1), colors.HexColor("#E8F5E9")),
                    ('GRID', (0,0), (-1,-1), 1, colors.black),
                ]))
                elements.append(table)
            
            elif report.type == "budget_analytics":
                total_budget = db.query(func.sum(Tender.budget)).filter(Tender.budget.isnot(None)).scalar() or 0
                active_tenders = db.query(Tender).filter(Tender.status == "active").count()
                closed_tenders = db.query(Tender).filter(Tender.status == "awarded").count()
                
                data = [
                    ["Метрика", "Значение"],
                    ["Жалпы бюджет", f"{total_budget} ₸"],
                    ["Активті тендерлер", str(active_tenders)],
                    ["Аяқталған тендерлер", str(closed_tenders)]
                ]
                
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#2196F3")),
                    ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0,0), (-1,0), 12),
                    ('BACKGROUND', (0,1), (-1,-1), colors.HexColor("#E3F2FD")),
                    ('GRID', (0,0), (-1,-1), 1, colors.black),
                ]))
                elements.append(table)

            # Save PDF
            doc.build(elements)
            result_file = pdf_path
            file_type = "pdf"

        elif report.type in ["supplier_ratings_excel", "supplier_performance"]:
            xlsx_filename = f"report_{report_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
            xlsx_path = os.path.join(UPLOAD_DIR, xlsx_filename)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Есеп"
            
            if report.type == "supplier_ratings_excel":
                ws.append(["ID", "Жеткізуші", "Сапасы", "Жеткізуі", "Көмік", "Бағасы", "Орташа рейтинг"])
                ratings = db.query(SupplierRating).all()
                for rating in ratings:
                    avg = (rating.quality_score + rating.delivery_score + rating.communication_score + rating.price_score) / 4
                    ws.append([rating.id, rating.supplier.full_name, rating.quality_score, rating.delivery_score, rating.communication_score, rating.price_score, round(float(avg), 2)])
                
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="#FF9800", end_color="#FF9800", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
            
            elif report.type == "supplier_performance":
                ws.append(["ID", "Жеткізуші", "Қатысқан тендерлер", "Жіберген ұсыныстар", "Орташа рейтинг"])
                # Simplified for now
                ratings = db.query(SupplierRating).all()
                for rating in ratings:
                    avg = (rating.quality_score + rating.delivery_score + rating.communication_score + rating.price_score) / 4
                    ws.append([rating.id, rating.supplier.full_name, 0, 0, round(float(avg), 2)])
                
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="#9C27B0", end_color="#9C27B0", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

            wb.save(xlsx_path)
            result_file = xlsx_path
            file_type = "xlsx"

        # Update report
        report.file_path = result_file
        report.file_type = file_type
        report.status = ReportStatus.COMPLETED
        db.commit()
        
        return {"status": "success", "file_path": result_file, "file_type": file_type}
    except Exception as e:
        report.status = ReportStatus.FAILED
        db.commit()
        return {"status": "error", "message": str(e)}
    finally:
        db.close()
